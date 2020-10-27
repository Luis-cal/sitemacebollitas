from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.http import JsonResponse
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import ListView, CreateView, UpdateView, DeleteView,TemplateView
from django.http.response import HttpResponse
from django.shortcuts import render
from core.erp.forms import ProductForm
from core.erp.mixins import ValidatePermissionRequiredMixin
from core.erp.models import Product
from openpyxl import Workbook
from tablib import Dataset
from core.erp.admin import ProductResource,SaleResource
#Nos devuelve un objeto resultado, en este caso un archivo de excel
def importar(request):  
   
   if request.method == 'POST':  
     persona_resource = ProductResource()  
     dataset = Dataset()  
     #print(dataset)  
     nuevas_personas = request.FILES['xlsfile']  
     #print(nuevas_personas)  
     imported_data = dataset.load(nuevas_personas.read())  
     #print(dataset)  
     result = persona_resource.import_data(dataset, dry_run=True) # Test the data import  
     #print(result.has_errors())  
     if not result.has_errors():  
       persona_resource.import_data(dataset, dry_run=False) # Actually import now 

   success_url= reverse_lazy('erp:product_list') 
   return render(request, 'importar.html') 
def impo(request):  
   
   if request.method == 'POST':  
     persona_resource = ProductResource()  
     dataset = Dataset()  
     #print(dataset)  
     nuevas_personas = request.FILES['xlsfile']  
     #print(nuevas_personas)  
     imported_data = dataset.load(nuevas_personas.read())  
     #print(dataset)  
     result = persona_resource.import_data(dataset, dry_run=True) # Test the data import  
     #print(result.has_errors())  
     if not result.has_errors():  
       persona_resource.import_data(dataset, dry_run=False) # Actually import now 

   success_url= reverse_lazy('erp:sale_list') 
   return render(request, 'impo.html') 






class ProductListView(LoginRequiredMixin, ValidatePermissionRequiredMixin, ListView):
    model = Product
    template_name = 'product/list.html'
    permission_required = 'view_product', 'change_product', 'delete_product', 'add_product'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            action = request.POST['action']
            if action == 'searchdata':
                data = []
                for i in Product.objects.all():
                    data.append(i.toJSON())
                    print(data[-1])
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado de Productos'
        context['create_url'] = reverse_lazy('erp:product_create')
        context['list_url'] = reverse_lazy('erp:product_list')
        context['entity'] = 'Productos'
        return context


class ProductCreateView(LoginRequiredMixin, ValidatePermissionRequiredMixin, CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'product/create.html'
    success_url = reverse_lazy('erp:product_list')
    permission_required = 'add_product'
    url_redirect = success_url

    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'add':
                form = self.get_form()
                data = form.save()
            else:
                data['error'] = 'No ha ingresado a ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Creación de un Producto'
        context['entity'] = 'Productos'
        context['list_url'] = self.success_url
        context['action'] = 'add'
        return context


class ProductUpdateView(LoginRequiredMixin, ValidatePermissionRequiredMixin, UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'product/create.html'
    success_url = reverse_lazy('erp:product_list')
    permission_required = 'change_product'
    url_redirect = success_url

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'edit':
                form = self.get_form()
                data = form.save()
            else:
                data['error'] = 'No ha ingresado a ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Edición de un Producto'
        context['entity'] = 'Productos'
        context['list_url'] = self.success_url
        context['action'] = 'edit'
        return context


class ProductDeleteView(LoginRequiredMixin, ValidatePermissionRequiredMixin, DeleteView):
    model = Product
    template_name = 'product/delete.html'
    success_url = reverse_lazy('erp:product_list')
    permission_required = 'delete_product'
    url_redirect = success_url

    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            self.object.delete()
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Eliminación de un Producto'
        context['entity'] = 'Productos'
        context['list_url'] = self.success_url
        return context

    
class ReportePersonasExcel(TemplateView):
     
     
    def get(self, request, *args, **kwargs):
        personas = Product.objects.all()
        
        wb = Workbook()
        
        ws = wb.active
        
        ws['A1'] = 'TABULADOR DE PRECIOS'
        
        ws.merge_cells('A1:E1')
        
        ws['A3'] = 'ID'
        ws['B3'] = 'NOMBRE'
        ws['C3'] = 'UNIDAD'
        ws['D3'] = 'CATEGORIA'
        ws['E3'] = 'PRECIO UNITARIO'
        ws['F3'] = 'IVA'
        ws['G3'] = 'PU+IVA'       
        cont=7
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for persona in personas:
            ws.cell(row=cont,column=1).value = persona.id
            ws.cell(row=cont,column=2).value = persona.name
            ws.cell(row=cont,column=3).value = persona.unidad
            ws.cell(row=cont,column=4).value = persona.cat
            ws.cell(row=cont,column=5).value = persona.pvp
            ws.cell(row=cont,column=6).value = persona.iva
            ws.cell(row=cont,column=7).value = persona.pui
            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="TABULADOR PRODUCTOS Excel.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
        

      


